"""Classes and functions for scraping www.belastingdienst.nl (version 2.7).

Classes in this module:

- ScrapeDB: encapsulation of an SQLite scrape database
- DataSheet: spreadsheet object to save large amounts of tabular data

Functions in this module:

- setup_file_logging: enable uniform logging for all modules
- scrape_page: scrape an html page and create an bs4 representation of it
- valid_path: validate that path can be used for an html scrape
- page_links: retrieve all links from the body of a page
- page_text: retrieve essential text content from a page

Module public constants:

- dv_types, bib_types, alg_types: sets of pagestypes that are considered to
    belong to a specific page category
"""

import re
import copy
from pathlib import Path
import logging
from typing import Dict, Union
import requests
import sqlite3
import zlib
import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.cell import WriteOnlyCell
from datetime import date

dv_types = {'bld-filter', 'bld-dv-content'}
bib_types = {'bld-bd', 'bld-cluster', 'bld-direction', 'bld-landing',
             'bld-overview', 'bld-sitemap', 'bld-target',
             'bld-targetGroup',
             'bld-concept', 'bld-faq'}
alg_types = {'bld-outage', 'bld-newsItem', 'bld-iahWrapper'}

_re_domain = re.compile(r'^https?://([\w-]*\.)*[\w-]*(?=/)')
_re_path = re.compile(r'^/[^/]')
_re_network_path = re.compile(r'^//[^/]')
_re_protocol = re.compile(r'^[a-z]{3,6}:')


class ScrapeDB:
    """Class encapsulating a scrape database.

    All actions on the scrape database are handled via the class methods.

    Class constants define some of the class behaviour.
    """

    version = '2.4'
    extracted_fields = [
        ('title', 'TEXT'),
        ('num_h1s', 'INTEGER'),
        ('first_h1', 'TEXT'),
        ('language', 'TEXT'),
        ('modified', 'DATE'),
        ('pagetype', 'TEXT'),
        ('classes', 'TEXT')
    ]
    derived_fields = [
        ('business', 'TEXT'),
        ('category', 'TEXT')
    ]

    def __init__(self, db_file, create=False):
        """Initiates the database object that encapsulates a scrape database.

        Writes the database version in the parameters table while creating a
        database. Reports an error if a database is opened with an incompatible
        version.

        Args:
            db_file (Path): name or path of the database file
            create (bool): create & connect database if True, else just connect
        """
        #
        # When db is on a networked drive maybe use next SQLite options to
        # improve query speed:
        #     PRAGMA synchronous = OFF
        #     PRAGMA journal_mode = PERSIST
        #
        self.db_file = db_file
        self.db_con = sqlite3.connect(self.db_file, isolation_level=None)
        self.exe = self.db_con.execute
        if create:
            self.exe('''
                CREATE TABLE pages (
                    page_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    path    TEXT NOT NULL UNIQUE,
                    doc     BLOB NOT NULL)''')
            self.exe('CREATE UNIQUE INDEX idx_pages_path ON pages (path)')
            self.exe('''
                CREATE TABLE redirs (
                    req_path   TEXT PRIMARY KEY NOT NULL UNIQUE,
                    redir_path TEXT NOT NULL,
                    type       TEXT)''')
            self.exe('''
                CREATE TABLE links (
                    page_id	  INTEGER NOT NULL,
                    link_text TEXT,
                    link_id   INTEGER,
                    ext_url   TEXT,
                    FOREIGN KEY (page_id, link_id)
                    REFERENCES pages (page_id, page_id)
                        ON UPDATE RESTRICT
                        ON DELETE RESTRICT)''')
            self.exe('''
                CREATE VIEW "links_explicit" AS
                    SELECT
                        l.page_id, p1.path AS page_path, l.link_text,
                        l.link_id, p2.path AS link_path, ext_url
                    FROM links AS l
                        JOIN pages AS p1 USING (page_id) 
                        LEFT JOIN pages AS p2 ON link_id = p2.page_id''')
            self.exe('''
                CREATE TABLE parameters (
                    name  TEXT PRIMARY KEY NOT NULL UNIQUE,
                    value TEXT NOT NULL)''')
            self.exe(
                f'INSERT INTO parameters VALUES ("db_version", {self.version})')
            logging.info(f'New scrape.db v{self.version} created')
        else:
            qry = 'SELECT value FROM parameters WHERE name = "db_version"'
            db_version = self.exe(qry).fetchone()[0]
            if db_version != self.version:
                logging.error(f'Incompatible database version: {db_version}')
                raise sqlite3.DatabaseError(
                    f'Incompatible database version: {db_version}')

    def close(self):
        """Close the connection to the database.

        Returns:
            None
        """
        self.db_con.close()

    def add_page(self, path, doc):
        """Add a scraped page.

        Args:
            path (str): path relative to the root of the scrape
            doc (str): complete scraped html of the page

        Returns:
            int|None: id of the page if not yet saved, None otherwise
        """
        qry = 'INSERT INTO pages (path, doc) VALUES (?, ?)'
        try:
            return self.exe(qry, [path, zlib.compress(doc.encode())]).lastrowid
        except sqlite3.IntegrityError:
            return None

    def get_page(self, path):
        """Get the id and complete doc string of a page.

        Args:
            path (str): path of the page, relative to root of scrape

        Returns:
            (int, str)|None: tuple (page_id, doc_string) or None if no match
        """
        qry = 'SELECT page_id, doc FROM pages WHERE path = ?'
        page = self.exe(qry, [path]).fetchone()
        if page:
            return page[0], zlib.decompress(page[1]).decode()
        else:
            return None

    def pages(self):
        """Generator for all pages of a stored scrape.

        Iterates over the pages stored in the database, yielding a threefold
        tuple containing: the id of the page, the page path (relative to the
        root of the scrape) and the scraped page as string.

        Returns:
            (int, str, str): page_id, page_path, page_string
        """
        qry = 'SELECT page_id, path, doc FROM pages'
        for page_id, path, doc in self.exe(qry):
            yield page_id, path, zlib.decompress(doc).decode()

    def num_pages(self):
        """Get total number of pages.

        Returns:
            int: number of pages
        """
        return self.exe('SELECT count(*) FROM pages').fetchone()[0]

    def add_redir(self, req_path, redir_path, redir_type):
        """Add a redirect that occurred during a page scrape.

        Args:
            req_path (str): requested path
            redir_path (str): path of the redirect
            redir_type (int|str): characterisation of the redirect: status code
                or textual

        Returns:
            None
        """
        qry = 'INSERT INTO redirs (req_path, redir_path, type) VALUES (?, ?, ?)'
        try:
            self.exe(qry, [req_path, redir_path, redir_type])
        except sqlite3.IntegrityError:
            pass
        return None

    def redirs(self):
        """Generator for all redirects of a stored scrape.

        Iterates over the redirects stored in the database, yielding a
        threefold tuple containing: the requested path, the path and the type
        of the redirect. All paths are relative to the root_url of the scrape.

        Returns:
            (str, str, str): requested path, path of redirect, type of redirect
        """
        qry = 'SELECT req_path, redir_path, type FROM redirs'
        for req_path, redir_path, redir_type in self.exe(qry):
            yield req_path, redir_path, redir_type

    def get_def_url(self, req_path):
        """Get the definitive url or path of a page.

        This definitive url or path is deduced via the available redirects in
        the redirs table. In case no redirect is available in the redirs
        table for the requested path, it will be returned unaltered.

        Args:
            req_path (str): requested path relative to root_url

        Returns:
            str: final redirected url or path relative to root_url
        """
        qry = 'SELECT redir_path, type FROM redirs WHERE req_path = ?'
        path = req_path
        while True:
            redir = self.exe(qry, [path]).fetchone()
            if redir:
                path, redir_type = redir
                if redir_type == 'alias':
                    # an alias redirects to the definitive path
                    redir = self.exe(qry, [path]).fetchone()
                    # test if definitive path does not get redirected itself
                    if redir:
                        logging.warning(
                            f'Definitive path gets redirected: {path}')
                    return path
                else:
                    # no alias redir, so maybe still another redir to go
                    continue
            else:
                return path

    def upd_par(self, name, value):
        """Insert or update a parameter.

        Args:
            name (str): name of the parameter
            value (str|int|float): value of the parameter

        Returns:
            None
        """
        qry = 'INSERT OR REPLACE INTO parameters (name, value) VALUES (?, ?)'
        self.exe(qry, [name, value])

    def get_par(self, name):
        """Get the value of a parameter.

        Args:
            name (str): name of the parameter

        Returns:
            str|int|float: value of the parameter
        """
        qry = 'SELECT value FROM parameters WHERE name = ?'
        result = self.exe(qry, [name]).fetchone()
        if not result:
            return None
        value = result[0]
        try:
            value = int(value)
        except ValueError:
            try:
                value = float(value)
            except ValueError:
                pass
        return value

    def fetch_pages_links(self):
        """Repopulates the links table with the relevant links of all pages.

        This method will take quite some time to complete. With a typical
        number or 150.000 links and depending on available computer
        resources, it can run for a quarter of an hour upto more than one hour.
        """

        # purge links table
        self.exe('DELETE FROM links')

        # (re)populate links table
        num_pages = self.num_pages()
        root_url = self.get_par('root_url')
        timestamp = self.get_par('timestamp')
        start_time = time.time()

        logging.info('Populating links table started')

        # cycle over all pages
        page_num = 0
        for page_id, page_path, page_string in self.pages():
            page_num += 1
            soup = BeautifulSoup(page_string, features='lxml')
            links = page_links(soup, root_url, root_rel=True,
                               excl_hdr_ftr=True)

            # cycle over all links of this page
            for link_text, link_url in links:
                def_url = self.get_def_url(link_url)
                page = self.get_page(def_url)
                if page:
                    # the link points to an internal page
                    link_id = page[0]
                    link_url = None
                else:
                    # because the link destination is not in the pages table,
                    # it is considered external
                    link_id = None
                self.exe('''
                    INSERT INTO links (page_id, link_text, link_id, ext_url)
                    VALUES (?, ?, ?, ?)''',
                         [page_id, link_text, link_id, link_url])

            # print progress and prognosis
            if page_num % 100 == 0:
                page_time = (time.time() - start_time) / page_num
                togo_time = int((num_pages - page_num) * page_time)
                print(f'fetching links from pages of {timestamp} - togo: '
                      f'{num_pages - page_num} pages / '
                      f'{togo_time // 60}:{togo_time % 60:02} min')

        logging.info('Populating links table completed\n')

    def links(self):
        """Generator for all links of a stored scrape.

        Iterates over the links stored in the database, yielding a fourfold
        tuple:

        - path of the originating page (relative to the root of the scrape)
        - text of the link
        - if internal relative to scrape root, path of the link, else None
        - if external relative to scrape root, full url of the link, else None

        Returns:
            (str, str, str|None, str|None): (page path, link text, link path,
                link url)
        """
        qry = 'SELECT page_path, link_text, link_path, ext_url FROM links_expl'
        for page_path, link_text, link_path, ext_url in self.exe(qry):
            yield page_path, link_text, link_path, ext_url

    def page_full_info(self, path):
        """Get all available information of a page.

        The returned dictionary has the next contents:

            - 'page_id': (int) page_id
            - 'path': (str) path
            - 'doc': (str) html source
            - 'title': (str) title
            - 'num_h1s': (int) number of h1 tags
            - 'first_h1': (str) text of the first h1 tag
            - 'language': (str) language
            - 'modified': (date) last modification date
            - 'pagetype': (str) type
            - 'classes': (str) classes separated by spaces
            - 'business': (str) 'belastingen', 'toeslagen' or 'douane'
            - 'category': (str) 'dv', 'bib' or 'alg'

        Args:
            path (str): path of the page

        Returns:
            dictionary[str, str|date|None] | None: info name:value pair
        """
        qry = "PRAGMA table_info('pages_full')"
        fields = [r[1] for r in self.exe(qry).fetchall()]
        qry = "SELECT * FROM pages_full WHERE path = ?"
        row = self.exe(qry, [path]).fetchone()
        if row:
            # type hint to prohibit warnings
            info: Dict[str, Union[date, str, bytes, None]]
            info = dict(zip(fields, row))
            mdate = info['modified']
            info['modified'] = date.fromisoformat(mdate) if mdate else None
            info['doc'] = zlib.decompress(info['doc']).decode()
            return info
        else:
            return None

    def pages_full(self):
        """Page generator yielding all available information per page.

        The yielded dictionary has the next contents:

            - 'page_id': (int) page_id
            - 'path': (str) path
            - 'doc': (str) html source
            - 'title': (str) title
            - 'num_h1s': (int) number of h1 tags
            - 'first_h1': (str) text of the first h1 tag
            - 'language': (str) language
            - 'modified': (date) last modification date
            - 'pagetype': (str) type
            - 'classes': (str) classes separated by spaces
            - 'business': (str) 'belastingen', 'toeslagen' or 'douane'
            - 'category': (str) 'dv', 'bib' or 'alg'

        Yields:
            dictionary[str, str|date|None]: info name:value pair
        """
        qry = "PRAGMA table_info('pages_full')"
        fields = [r[1] for r in self.exe(qry).fetchall()]
        qry = "SELECT * FROM pages_full"
        for row in self.exe(qry):
            # type hint to prohibit warnings
            info: Dict[str, Union[date, str, bytes, None]]
            info = dict(zip(fields, row))
            mdate = info['modified']
            info['modified'] = date.fromisoformat(mdate) if mdate else None
            info['doc'] = zlib.decompress(info['doc']).decode()
            yield info

    def extract_pages_info(self):
        """Add table with information extracted from all pages.

        Extracted information concerns data that is available within the page
        which is stored seperately in the pages table. Storing this data in a
        seperate table is strictly redundant, but serves faster access.

        The fields of the pages_info table are defined by the class constants
        extracted_fields and derived_fields. Besides the pages_info table a
        pages_full view is added that joins the pages table with the
        pages_info table. Existing table and/or view are deleted before
        creating new ones.

        The following information is added for each page:

        - title: content of <title> tag
        - num_h1s: number <h1>'s
        - first_h1: text of the first h1
        - language: content of <meta name="language" content="xx" />
        - modified: content of <meta name="DCTERMS.modified" content="date" />
        - pagetype: attribute value of <body data-pageType="...">
        - classes: attribute value of <body class="...">

        The pages_info table accommodates additional fields to contain
        derived information for each page. This is further detailed in the
        derive_pages_info method of this class.

        It will be logged when tags or attributes are missing or values are
        invalid.
        """

        # create new pages_info table
        self.exe('DROP TABLE IF EXISTS pages_info')
        fields = self.extracted_fields + self.derived_fields
        info_columns = ', '.join([f'{f[0]} {f[1]}' for f in fields])
        self.exe(f'''
            CREATE TABLE pages_info (
                page_id	 INTEGER PRIMARY KEY NOT NULL UNIQUE,
                {info_columns},
                FOREIGN KEY (page_id)
                REFERENCES pages (page_id)
                    ON UPDATE RESTRICT
                    ON DELETE RESTRICT)''')

        # create new pages_full view
        self.exe('DROP VIEW IF EXISTS pages_full')
        self.exe('''
            CREATE VIEW pages_full AS
                SELECT *
                FROM pages
                LEFT JOIN pages_info USING (page_id)''')
        self.exe('VACUUM')
        logging.info(
            'Pages_info table and pages_full view (re)created in scrape.db')

        # extract info from all pages while populating the pages_info table
        num_pages = self.num_pages()
        timestamp = self.get_par('timestamp')
        start_time = time.time()

        logging.info('Extracting info from pages started')

        # cycle over all pages
        page_num = 0
        for page_id, path, page_string in self.pages():
            page_num += 1
            soup = BeautifulSoup(page_string, features='lxml')
            info = {'page_id': page_id}

            # get title
            title = soup.head.title
            if not title:
                logging.warning(f'Page has no <title> tag: {path}')
                title = None
            else:
                title = title.text
                if not title:
                    logging.warning(f'Page with empty title: {path}')
            info['title'] = title

            # get language
            language = soup.head.find('meta', attrs={'name': 'language'})
            if not language:
                logging.warning(
                    f'Page has no <meta name="language"/> tag: {path}')
                language = None
            else:
                language = language['content']
                if not language:
                    logging.warning(f'Page with empy language: {path}')
            info['language'] = language

            # get date modified
            modified = soup.head.find('meta',
                                      attrs={'name': 'DCTERMS.modified'})
            if not modified:
                logging.warning(
                    f'Page has no tag <meta name="DCTERMS.modified"/>: {path}')
                modified = None
            else:
                try:
                    modified = date.fromisoformat(modified['content'])
                except ValueError:
                    logging.warning(
                        f'Page with improper modification date: {path}')
                    modified = None
            info['modified'] = modified

            # get type of page
            if 'data-pagetype' not in soup.body.attrs:
                logging.warning('Page has no data-pagetype attribute in the '
                                f'<body> tag: {path}')
                pagetype = None
            else:
                pagetype = soup.body['data-pagetype']
                if not pagetype:
                    logging.warning(
                        f'Page with empty pagetype in <body> tag: {path}')
            info['pagetype'] = pagetype

            # get classes
            if 'class' not in soup.body.attrs:
                logging.warning(
                    f'Page has no class attribute in the <body> tag: {path}')
                classes = None
            else:
                classes = soup.body['class']
                if not classes:
                    logging.warning(
                        f'Page with empty class in <body> tag: {path}')
            info['classes'] = ' '.join(classes) if classes else None

            # get info from <h1> tags
            h1s = []
            for h1 in soup.find_all('h1'):
                h1s.append(h1.text)
            if len(h1s) == 0:
                logging.warning(f'Page without h1: {path}')
            info['num_h1s'] = len(h1s)
            info['first_h1'] = h1s[0] if h1s else None

            # add info to the database
            self.exe('''
                INSERT INTO pages_info
                    (page_id, title, num_h1s, first_h1,
                    language, modified, pagetype, classes)
                VALUES
                    (:page_id, :title, :num_h1s, :first_h1,
                    :language, :modified, :pagetype, :classes)''',
                     info)

            # print progress and prognosis
            if page_num % 250 == 0:
                page_time = (time.time() - start_time) / page_num
                togo_time = int((num_pages - page_num) * page_time)
                print(
                    f'adding extracted info to scrape database of {timestamp} '
                    f'- togo: {num_pages - page_num} pages / '
                    f'{togo_time // 60}:{togo_time % 60:02} min')

        logging.info('Extracting info from pages completed\n')

    def derive_pages_info(self):
        """Add table with derived information for all pages.

        Derived information as such is not available within a page,
        but calculated or interpreted from other information. To derive this
        information, the extracted information should already be available in
        the pages_info table. This can be accomplished by using the
        extract_pages_info method of this class.

        The fields in which the derived info is saved, are already available in
        the pages_info table. In case a field is added, the extract_pages_info
        method can be used to recreate the pages_table, while adding the
        extra fields. The class constants extracted_fields and derived_fields
        define together the fields that are created in the pages_info table.

        The following information is added for each page:

        - business: 'belastingen', 'toeslagen' or 'douane'
        - category: 'dv', 'bib' or 'alg'

        It will be logged when info can not be derived due to inconsistent or
        unavailable information.
        """

        # clear derived info fields in pages_info table
        set_cols = ', '.join([f'{f[0]} = NULL' for f in self.derived_fields])
        self.exe(f'UPDATE pages_info SET {set_cols}')

        # derive info from all pages while updating the pages_info table
        num_pages = self.num_pages()
        timestamp = self.get_par('timestamp')
        start_time = time.time()

        logging.info('Deriving info from pages started')

        # cycle over all pages, with wrapper pages after all others, because
        # the wrapper category is determined by the catogories of pages
        # linking to that wrapper page
        for_qry = '''
            SELECT page_id, pagetype, classes
            FROM pages_info
            ORDER BY
                CASE pagetype
                    WHEN 'bld-wrapper' THEN 2 ELSE 1
                END'''
        page_num = 0
        for page_id, pagetype, classes in self.exe(for_qry).fetchall():
            page_num += 1
            fields = {'page_id': page_id}

            # determine business
            if classes:
                if 'toeslagen' in classes:
                    business = 'toeslagen'
                elif 'douane' in classes:
                    business = 'douane'
                else:
                    business = 'belastingen'
            else:
                business = None
            fields['business'] = business

            # determine category: dv, bib or alg
            if pagetype in dv_types:
                category = 'dv'
            elif pagetype in bib_types:
                category = 'bib'
            elif pagetype in alg_types:
                category = 'alg'
            elif pagetype == 'bld-wrapper':
                cat_qry = '''
                    SELECT category
                    FROM links
                        JOIN pages_info USING (page_id)
                    WHERE link_id = ?
                    GROUP BY category'''
                categories = self.exe(cat_qry, [page_id]).fetchall()
                if len(categories) == 1:
                    # all pages linking to the wrapper are of the same category
                    category = categories[0][0]
                else:
                    category = 'alg'
            else:
                category = 'unknown'
            fields['category'] = category

            # save the derived fields in the pages_info table
            self.exe('''
                UPDATE pages_info
                SET business = :business,
                    category = :category
                WHERE
                    page_id = :page_id''', fields)

            # print progress and prognosis
            if page_num % 500 == 0:
                page_time = (time.time() - start_time) / page_num
                togo_time = int((num_pages - page_num) * page_time)
                print(
                    f'adding derived info to scrape database of {timestamp} - '
                    f'togo: {num_pages - page_num} pages / '
                    f'{togo_time // 60}:{togo_time % 60:02} min')

        logging.info('Deriving info from pages completed\n')


class DataSheet:
    """Spreadsheet object to save large amounts of tabular data.

    Basically an encapsulation of an OpenPyXL Workbook object to circumvent some
    limitations of a write-only workbook. This optimized write-only mode of
    OpenPyXL is needed to limit resource usage while saving virtually unlimited
    amounts of data.
    """

    def __init__(self, sheet_name, *col_spec):
        """Create a workbook with one sheet to receive many rows of data.

        The sheet that is created in the workbook will have a formatted first
        row with column titles and column widths as specified. Width units are
        roughly equivalent to number of characters (the real story is quite
        complex).
        The workbook will be written to disk upon calling the save method.

        Args:
            sheet_name (str): tab-name of the sheet
            col_spec ((str, int)): title and width of column
        """
        self.wb = Workbook(write_only=True)
        self.ws = self.wb.create_sheet(sheet_name)
        self.ws.freeze_panes = 'A2'
        self.ws.sheet_view.zoomScale = 90
        col = 0
        title_cells = []
        for col_title, col_width in col_spec:
            col += 1
            self.ws.column_dimensions[self.col_ref(col)].width = col_width
            cell = WriteOnlyCell(self.ws, value=col_title)
            cell.font = Font(name='Calibri', bold=True, color='FFFFFF')
            cell.fill = PatternFill(fgColor='808080', fill_type='solid')
            title_cells.append(cell)
        self.ws.append(title_cells)
        self.last_row = 1
        self.num_cols = col

    def append(self, values):
        """Append a group of values at the bottom of the sheet.

        Args:
            values (list|tuple|range|dict|generator):
                iterable containing values to append

        Returns:
            None
        """
        if len(values) > self.num_cols:
            raise ValueError('Too much values for row.')
        self.ws.append(values)
        self.last_row += 1

    def save(self, filename):
        """Save the workbook.

        Args:
            filename (str): name for the worksheet file

        Returns:
            None
        """
        last_cell = self.col_ref(self.num_cols) + str(self.last_row)
        self.ws.auto_filter.ref = 'A1:' + last_cell
        self.wb.save(filename)

    @staticmethod
    def col_ref(num):
        """Convert column number to capitals index.

        Column number should be between 1 and 702 inclusive.

        Args:
            num (int): column number

        Returns:
            str: capitals index
        """
        num -= 1
        if num < 0:
            raise ValueError('Column number less than 1')
        elif num < 26:
            return chr(num % 26 + 65)
        elif num < 702:
            return chr(num // 26 + 64) + chr(num % 26 + 65)
        else:
            raise ValueError('Column number cannot be greater than 702.')


def setup_file_logging(directory, log_level=logging.INFO):
    """Enable uniform logging for all modules.

    Args:
        directory (Path): path of the directory of the logfile
        log_level (int): the lowest severity level that will be logged
            symbolic values are available from the logging module:
            CRITICAL, ERROR, WARNING, INFO, DEBUG, NOTSET

    Returns:
        None
    """
    logging.basicConfig(
        filename=str(directory / 'log.txt'),
        format='[%(asctime)s] %(levelname)-8s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
        level=log_level,
        force=True)


def valid_path(path):
    """Validate that url can be used for an html scrape.

    The function tests against the next exceptions:
        - ends with .xml

    Args:
        path (str): path to be checked

    Returns:
        bool: the result of the validation
    """
    if path.endswith('.xml'):
        logging.debug('Path ending in .xml: %s' % path)
        return False
    else:
        return True


def scrape_page(root_url, req_url):
    """Scrape an html page.

    Since there can be more than one redirect per requested page, the last
    element in the return tuple is a list of redirects instead of a single
    redirect (a 301 is often followed by a 302).
    The first item in the returned tuple is the url that comes from the content
    attribute of the <meta name="DCTERMS.identifier"> tag. This is the
    definitive url as generated by the WCM system.
    All url's are absolute.

    Args:
        root_url (str): url that will be treated as the base of the scrape;
            links starting with root_url are interpreted as within scope
        req_url (str): url of requested page

    Returns:
        (str, BeautifulSoup, str, list of (str, str, str)):
            definitive url of the page,
            bs4 representation of the page,
            complete page as string,
            list of (requested url, url of the response, type of redirect)
    """
    redirs = []

    while True:
        # cycle until no rewrites or redirects
        resp = requests.get(req_url)
        if resp.status_code != 200:
            logging.error(f'Unexpected response from {req_url}; '
                          f'status code is {resp.status_code}.')
            raise requests.RequestException

        # read and parse the response into a soup document
        # resp_url = resp.url
        page_as_string = resp.text
        soup = BeautifulSoup(page_as_string, features='lxml')

        # are there any redirects?
        if len(resp.history) != 0:
            i_req_url = req_url
            for i_resp in resp.history:
                if i_req_url.startswith(root_url):
                    # requested page is within scope of scrape
                    i_resp_url = i_resp.headers['Location']
                    redirs.append((i_req_url, i_resp_url, i_resp.status_code))
                    if i_resp.status_code not in (301, 302):
                        # just to know if this occurs; probably not
                        logging.warning(
                            f'Requesting {i_req_url} responded with '
                            f'status code {i_resp.status_code}.')
                    i_req_url = i_resp_url

        # do we have a client-side redirect page via the next header tag?
        #       <meta http-equiv="refresh" content="0;url=...">
        meta_tag = soup.head.find('meta', attrs={'http-equiv': 'refresh'})
        if meta_tag:
            resp_url = meta_tag['content'].partition('url=')[2]
            # complete url if necessary
            if re.match(_re_path, resp_url):
                domain = _re_domain.match(root_url)[0]
                resp_url = domain + resp_url
            elif re.match(_re_network_path, resp_url):
                protocol = _re_protocol.match(root_url)[0]
                resp_url = protocol + resp_url
            redirs.append((req_url, resp_url, 'client'))
            req_url = resp_url
            continue

        resp_url = resp.url
        if resp_url.startswith(root_url):
            # response url is within scope
            meta_tag = soup.head.find(
                'meta', attrs={'name': 'DCTERMS.identifier'})
            if meta_tag:
                def_url = meta_tag['content']
                # complete path
                if def_url.startswith('/wps/wcm/connect'):
                    domain = _re_domain.match(root_url)[0]
                    def_url = domain + def_url
                    if def_url != resp_url:
                        redirs.append((resp_url, def_url, 'alias'))
                else:
                    logging.warning('Non-standard definitive url; '
                                    f'falling back to: {resp_url}')
                    def_url = resp_url
            else:
                logging.error(
                    f'Page without definitive url; falling back to: {resp_url}')
                def_url = resp_url
        else:
            def_url = resp_url

        # return implicitly ends while loop
        return def_url, soup, page_as_string, redirs


def page_links(soup, root_url,
               root_rel=False, excl_hdr_ftr=True, remove_anchor=True):
    """Retrieve all links from the body of a page.

    The links will be absolute url's or paths relative to root_url.
    The returned links will be filtered according to the following criteria:

    - no links from <div id="bld-nojs">
    - no links from <header> or <footer> if excl_hdr_ftr == True
    - no links containing 'readspeaker' or 'adobe'
    - only links that start with /, // or protocol: (uri-scheme)

    Further processing of the links is determined by the arguments:

    - anchors are removed if remove_anchor == True
    - links are relative to root_url if root_rel == True, otherwise absolute

    Args:
        soup (BeautifulSoup): bs4 representation of the page
        root_url (str): the root url with which the page was scraped
        root_rel (bool): return links relative to root_url
        excl_hdr_ftr (bool): exclude links from header and footer branch
        remove_anchor (bool): remove anchors from links

    Returns:
        list of (str, str): list of (link text, link url) tuples
    """

    # make working copy of the doc, since removing branches is destructive
    sc = copy.copy(soup)

    # clear branches from which links are excluded
    div_nojs = sc.find('div', id='bld-nojs')
    if div_nojs:
        div_nojs.clear()
    if excl_hdr_ftr:
        if sc.body.header:
            sc.body.header.clear()
        if sc.body.footer:
            sc.body.footer.clear()

    # get links from remaining soup doc
    links = []
    for a_tag in sc.body.find_all('a', href=True):
        link = a_tag['href']
        if 'readspeaker' in link or 'adobe' in link:
            continue
        if remove_anchor:
            link = link.partition('#')[0]

        # make link into complete url's if necessary
        if re.match(_re_path, link):
            domain = _re_domain.match(root_url)[0]
            link = domain + link
        elif re.match(_re_network_path, link):
            protocol = _re_protocol.match(root_url)[0]
            link = protocol + link
        elif not re.match(_re_protocol, link):
            continue

        # make link relative to root_url if needed
        if root_rel:
            link = re.sub(root_url, '', link)

        if link:
            links.append((a_tag.text.strip(), link))

    # remove the working copy of the soup doc
    sc.decompose()
    return links


def page_text(soup):
    """Retrieve essential text content from a page.

    Only text from <body> is included.

    Next tags are excluded:

    - <header>
    - <footer>
    - <div id="bld-nojs">

    Leading and trailing whitespace is removed.

    Args:
        soup (BeautifulSoup): bs4 representation of a page

    Returns:
        str: newline separated chunks of extracted text
    """
    # make working copy of the soup doc, since removing branches is destructive
    sc = copy.copy(soup)

    # remove branches
    div_nojs = sc.find('div', id='bld-nojs')
    if div_nojs:
        div_nojs.clear()
    if sc.body.header:
        sc.body.header.clear()
    if sc.body.footer:
        sc.body.footer.clear()

    # get the text to return
    txt = sc.body.get_text(separator='\n', strip=True)

    # remove the working copy of the soup doc
    sc.decompose()
    return txt
